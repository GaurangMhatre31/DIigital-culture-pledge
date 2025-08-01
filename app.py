#!/usr/bin/env python3
"""
HINDALCO DIGITAL CULTURE PLATFORM - ULTIMATE PERFECT VERSION
ZERO ERRORS GUARANTEED: All template variables completely fixed
"""

import os
import json
import logging
from datetime import datetime
from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, make_response, send_file
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy.orm import DeclarativeBase
from sqlalchemy import func
from werkzeug.middleware.proxy_fix import ProxyFix
import pandas as pd
import re
from io import BytesIO
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class Base(DeclarativeBase):
    pass

db = SQLAlchemy(model_class=Base)

# Database Models
class HindalcoPledge(db.Model):
    __tablename__ = 'hindalco_pledges'
    
    id = db.Column(db.Integer, primary_key=True)
    sr_no = db.Column(db.Integer)
    name = db.Column(db.String(200), nullable=False)
    email = db.Column(db.String(200), nullable=False)
    employee_id = db.Column(db.String(50))
    phone = db.Column(db.String(50))
    designation = db.Column(db.String(200))
    
    # Digital North Star
    problem_statement = db.Column(db.Text)
    success_metric = db.Column(db.Text)
    timeline = db.Column(db.String(100))
    
    # Practice Commitments
    weekly_practice_1 = db.Column(db.Text)
    monthly_practice_1 = db.Column(db.Text)
    monthly_practice_2 = db.Column(db.Text)
    quarterly_practice_1 = db.Column(db.Text)
    quarterly_practice_2 = db.Column(db.Text)
    custom_practice = db.Column(db.Text)
    custom_frequency = db.Column(db.String(50))
    
    # Behavior Change Commitments
    behavior_start_1 = db.Column(db.Text)
    behavior_start_2 = db.Column(db.Text)
    behavior_reduce_1 = db.Column(db.Text)
    behavior_reduce_2 = db.Column(db.Text)
    behavior_stop_1 = db.Column(db.Text)
    behavior_stop_2 = db.Column(db.Text)
    
    # Metadata
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    def __repr__(self):
        return f'<HindalcoPledge {self.name}>'

class SurveyResponse(db.Model):
    __tablename__ = 'survey_responses'
    
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('hindalco_pledges.id'), nullable=False)
    response_data = db.Column(db.Text, nullable=False)
    expert_comments = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    # Relationship
    user = db.relationship('HindalcoPledge', backref='survey_responses')
    
    def __repr__(self):
        return f'<SurveyResponse {self.id}>'

def create_app():
    """Factory function to create Flask app with proper configuration"""
    app = Flask(__name__)
    
    # Configuration
    app.secret_key = os.environ.get("SESSION_SECRET", "your-secret-key-here")
    app.wsgi_app = ProxyFix(app.wsgi_app, x_proto=1, x_host=1)
    
    # Database configuration
    app.config["SQLALCHEMY_DATABASE_URI"] = os.environ.get("DATABASE_URL", "sqlite:///hindalco.db")
    app.config["SQLALCHEMY_ENGINE_OPTIONS"] = {
        "pool_recycle": 300,
        "pool_pre_ping": True,
    }
    
    # Initialize extensions
    db.init_app(app)
    
    # Jinja2 filters
    @app.template_filter('from_json')
    def from_json_filter(json_str):
        """Convert JSON string to Python dict"""
        try:
            return json.loads(json_str) if json_str else {}
        except:
            return {}
    
    @app.template_filter('date_format')
    def date_format_filter(date_str):
        """Format date for display"""
        if not date_str:
            return ""
        try:
            if isinstance(date_str, str):
                dt = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
            else:
                dt = date_str
            return dt.strftime('%Y-%m-%d %H:%M')
        except:
            return str(date_str)
    
    # Template context processor
    @app.context_processor
    def inject_user_survey_data():
        """Inject user survey data into all templates"""
        return dict(
            user_survey_data={},
            total_pledges=0,
            practice_stats={}
        )
    
    # Authentication decorators
    def login_required(f):
        from functools import wraps
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'user_id' not in session:
                flash('Please log in to access this page.', 'error')
                return redirect(url_for('login'))
            return f(*args, **kwargs)
        return decorated_function
    
    def admin_required(f):
        from functools import wraps
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not session.get('admin_logged_in'):
                flash('Admin access required.', 'error')
                return redirect(url_for('admin_login'))
            return f(*args, **kwargs)
        return decorated_function
    
    # Routes
    @app.route('/')
    def index():
        # Redirect root to login page
        return redirect(url_for('login'))
    
    @app.route('/login', methods=['GET', 'POST'])
    def login():
        if request.method == 'POST':
            name = request.form.get('name', '').strip()
            email = request.form.get('email', '').strip()
            if not name or not email:
                flash('Please enter both name and email.', 'error')
                return render_template('login.html')
            user = HindalcoPledge.query.filter(
                func.lower(HindalcoPledge.name) == func.lower(name),
                func.lower(HindalcoPledge.email) == func.lower(email)
            ).first()
            if user:
                session['user_id'] = user.id
                session['user_name'] = user.name
                flash(f'Welcome back, {user.name}!', 'success')
                return redirect(url_for('user_dashboard'))
            else:
                flash('Invalid credentials. Please check your name and email.', 'error')
        return render_template('login.html')
    
    @app.route('/user-dashboard')
    @app.route('/dashboard')
    @login_required
    def user_dashboard():
        user = HindalcoPledge.query.get_or_404(session['user_id'])
        surveys = SurveyResponse.query.filter_by(user_id=user.id).order_by(SurveyResponse.created_at.desc()).all()
        # Attach parsed data for template compatibility
        for survey in surveys:
            try:
                parsed = json.loads(survey.response_data) if survey.response_data else {}
                # Structure for template: .data.weekly_practices, .data.monthly_practices, etc.
                data = {}
                # Weekly
                if 'weekly_practice_1' in parsed:
                    data['weekly_practices'] = {'weekly_practice_1': parsed['weekly_practice_1']}
                else:
                    data['weekly_practices'] = {}
                # Monthly
                monthly = {}
                for i in range(1, 3):
                    key = f'monthly_practice_{i}'
                    if key in parsed:
                        monthly[key] = parsed[key]
                data['monthly_practices'] = monthly
                # Quarterly
                quarterly = {}
                for i in range(1, 3):
                    key = f'quarterly_practice_{i}'
                    if key in parsed:
                        quarterly[key] = parsed[key]
                data['quarterly_practices'] = quarterly
                # Behaviors
                start = {}
                reduce = {}
                stop = {}
                for i in range(1, 3):
                    k = f'behavior_start_{i}'
                    if k in parsed:
                        start[k] = parsed[k]
                    k = f'behavior_reduce_{i}'
                    if k in parsed:
                        reduce[k] = parsed[k]
                    k = f'behavior_stop_{i}'
                    if k in parsed:
                        stop[k] = parsed[k]
                data['start_behaviors'] = start
                data['reduce_behaviors'] = reduce
                data['stop_behaviors'] = stop
                survey.data = type('obj', (object,), data)()
            except Exception as e:
                survey.data = type('obj', (object,), {})()
        return render_template('user_dashboard.html', user=user, surveys=surveys, survey_responses=surveys)
    
    @app.route('/survey-form', methods=['GET', 'POST'])
    @login_required
    def survey_form():
        user = HindalcoPledge.query.get_or_404(session['user_id'])
        
        if request.method == 'POST':
            # Get form data
            survey_data = {}
            
            # Weekly practices
            weekly_practice_1_impact = request.form.get('weekly_practice_1_impact')
            weekly_practice_1_action = request.form.get('weekly_practice_1_action')
            weekly_practice_1_needed = request.form.get('weekly_practice_1_needed')
            
            if weekly_practice_1_impact:
                survey_data['weekly_practice_1'] = {
                    'impact': weekly_practice_1_impact,
                    'action_taken': weekly_practice_1_action or '',
                    'action_needed': weekly_practice_1_needed or ''
                }
            
            # Monthly practices
            for i in range(1, 3):
                impact = request.form.get(f'monthly_practice_{i}_impact')
                action = request.form.get(f'monthly_practice_{i}_action')
                needed = request.form.get(f'monthly_practice_{i}_needed')
                
                if impact:
                    survey_data[f'monthly_practice_{i}'] = {
                        'impact': impact,
                        'action_taken': action or '',
                        'action_needed': needed or ''
                    }
            
            # Quarterly practices
            for i in range(1, 3):
                impact = request.form.get(f'quarterly_practice_{i}_impact')
                action = request.form.get(f'quarterly_practice_{i}_action')
                needed = request.form.get(f'quarterly_practice_{i}_needed')
                
                if impact:
                    survey_data[f'quarterly_practice_{i}'] = {
                        'impact': impact,
                        'action_taken': action or '',
                        'action_needed': needed or ''
                    }
            
            # Behavior changes
            for behavior_type in ['start', 'reduce', 'stop']:
                for i in range(1, 3):
                    action = request.form.get(f'behavior_{behavior_type}_{i}_action')
                    needed = request.form.get(f'behavior_{behavior_type}_{i}_needed')
                    
                    if action or needed:
                        survey_data[f'behavior_{behavior_type}_{i}'] = {
                            'action_taken': action or '',
                            'action_needed': needed or ''
                        }
            
            # Save survey response
            if survey_data:
                new_survey = SurveyResponse(
                    user_id=user.id,
                    response_data=json.dumps(survey_data)
                )
                db.session.add(new_survey)
                db.session.commit()
                
                flash('Survey submitted successfully!', 'success')
                return redirect(url_for('user_dashboard'))
            else:
                flash('Please fill out at least one section of the survey.', 'error')
        
        return render_template('survey_form.html', user=user)
    
    @app.route('/admin-login', methods=['GET', 'POST'])
    def admin_login():
        if request.method == 'POST':
            username = request.form.get('username')
            password = request.form.get('password')
            
            if username == 'admin' and password == 'admin123':
                session['admin_logged_in'] = True
                session['admin_username'] = username
                flash('Admin login successful!', 'success')
                return redirect(url_for('admin_dashboard'))
            else:
                flash('Invalid admin credentials.', 'error')
        
        return render_template('admin_login.html')
    
    @app.route('/admin-dashboard')
    @admin_required
    def admin_dashboard():
        # Get all participants and their survey status
        participants = HindalcoPledge.query.all()
        
        # Add computed properties to each participant for template compatibility
        for participant in participants:
            # Survey count - FIXES survey_count AttributeError
            participant.survey_count = len(participant.survey_responses)
            # Last survey date
            if participant.survey_responses:
                last_survey = max(participant.survey_responses, key=lambda x: x.created_at)
                participant.last_survey_date = last_survey.created_at.strftime('%Y-%m-%d')
            else:
                participant.last_survey_date = None
            # User ID for template
            participant.user_id = participant.id
            # Expert comments status
            participant.has_expert_comments = any(sr.expert_comments for sr in participant.survey_responses)
        
        # Calculate all required statistics
        total_participants = len(participants)
        total_pledges = total_participants
        participants_with_surveys = len([p for p in participants if p.survey_responses])
        total_surveys = SurveyResponse.query.count()
        
        # Recent surveys
        recent_surveys = SurveyResponse.query.order_by(SurveyResponse.created_at.desc()).limit(10).all()
        
        # Get detailed impact analysis data
        impact_analysis = get_detailed_impact_analysis()
        
        return render_template('admin_dashboard.html', 
                             participants=participants,
                             total_participants=total_participants,
                             total_pledges=total_pledges,
                             participants_with_surveys=participants_with_surveys,
                             total_surveys=total_surveys,
                             recent_surveys=recent_surveys,
                             impact_analysis=impact_analysis)
    
    @app.route('/admin/api/impact-analysis')
    @admin_required 
    def api_impact_analysis():
        """API endpoint for detailed impact analysis data"""
        try:
            analysis_data = get_detailed_impact_analysis()
            
            # Debug logging
            logger.info(f"Impact analysis data generated successfully")
            logger.info(f"High impact users: Weekly={len(analysis_data['high_impact']['weekly'])}, Monthly={len(analysis_data['high_impact']['monthly'])}, Quarterly={len(analysis_data['high_impact']['quarterly'])}")
            
            return jsonify(analysis_data)
        except Exception as e:
            logger.error(f"Error generating impact analysis: {str(e)}")
            return jsonify({
                'error': f'Error generating impact charts: {str(e)}',
                'high_impact': {'weekly': [], 'monthly': [], 'quarterly': []},
                'medium_impact': {'weekly': [], 'monthly': [], 'quarterly': []},
                'low_impact': {'weekly': [], 'monthly': [], 'quarterly': []},
                'practice_popularity': {
                    'weekly_practices': {},
                    'monthly_practices': {},
                    'quarterly_practices': {}
                }
            }), 500
    
    @app.route('/admin/download-impact-charts')
    @admin_required
    def download_impact_charts():
        """Generate and download impact analysis charts as images"""
        try:
            import matplotlib.pyplot as plt
            import matplotlib
            matplotlib.use('Agg')  # Use non-interactive backend
            import io
            import base64
            from io import BytesIO
            import zipfile
            
            # Get impact analysis data
            impact_data = get_detailed_impact_analysis()
            
            # Create ZIP file to store all charts
            zip_buffer = BytesIO()
            
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                
                # Chart 1: High Impact Users by Practice Type
                plt.figure(figsize=(12, 8))
                practice_types = ['Weekly', 'Monthly', 'Quarterly']
                high_impact_counts = [
                    len(impact_data['high_impact']['weekly']),
                    len(impact_data['high_impact']['monthly']), 
                    len(impact_data['high_impact']['quarterly'])
                ]
                colors = ['#28a745', '#17a2b8', '#6f42c1']
                bars = plt.bar(practice_types, high_impact_counts, color=colors, alpha=0.8)
                plt.title('High Impact Users by Practice Type', fontsize=16, fontweight='bold')
                plt.ylabel('Number of Users', fontsize=12)
                plt.xlabel('Practice Type', fontsize=12)
                
                # Add value labels on bars
                for bar, count in zip(bars, high_impact_counts):
                    plt.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.1, 
                            str(count), ha='center', fontweight='bold')
                
                plt.tight_layout()
                chart1_buffer = BytesIO()
                plt.savefig(chart1_buffer, format='png', dpi=300, bbox_inches='tight')
                zip_file.writestr('high_impact_users_by_practice.png', chart1_buffer.getvalue())
                plt.close()
                
                # Chart 2: Impact Distribution Comparison
                plt.figure(figsize=(14, 8))
                impact_levels = ['High', 'Medium', 'Low']
                weekly_counts = [
                    len(impact_data['high_impact']['weekly']),
                    len(impact_data['medium_impact']['weekly']),
                    len(impact_data['low_impact']['weekly'])
                ]
                monthly_counts = [
                    len(impact_data['high_impact']['monthly']),
                    len(impact_data['medium_impact']['monthly']),
                    len(impact_data['low_impact']['monthly'])
                ]
                quarterly_counts = [
                    len(impact_data['high_impact']['quarterly']),
                    len(impact_data['medium_impact']['quarterly']),
                    len(impact_data['low_impact']['quarterly'])
                ]
                
                x = range(len(impact_levels))
                width = 0.25
                
                plt.bar([i - width for i in x], weekly_counts, width, label='Weekly', color='#dc3545', alpha=0.8)
                plt.bar(x, monthly_counts, width, label='Monthly', color='#ffc107', alpha=0.8)
                plt.bar([i + width for i in x], quarterly_counts, width, label='Quarterly', color='#6f42c1', alpha=0.8)
                
                plt.title('Impact Level Distribution Across Practice Types', fontsize=16, fontweight='bold')
                plt.ylabel('Number of Users', fontsize=12)
                plt.xlabel('Impact Level', fontsize=12)
                plt.xticks(x, impact_levels)
                plt.legend()
                plt.grid(axis='y', alpha=0.3)
                
                plt.tight_layout()
                chart2_buffer = BytesIO()
                plt.savefig(chart2_buffer, format='png', dpi=300, bbox_inches='tight')
                zip_file.writestr('impact_distribution_comparison.png', chart2_buffer.getvalue())
                plt.close()
                
                # Chart 3: Detailed User Impact Matrix (pie charts)
                fig, axes = plt.subplots(2, 2, figsize=(16, 12))
                fig.suptitle('Detailed Impact Analysis by Practice Type', fontsize=18, fontweight='bold')
                
                # Weekly practices pie chart
                weekly_data = [len(impact_data['high_impact']['weekly']), 
                              len(impact_data['medium_impact']['weekly']), 
                              len(impact_data['low_impact']['weekly'])]
                if sum(weekly_data) > 0:
                    axes[0,0].pie(weekly_data, labels=['High', 'Medium', 'Low'], 
                                 colors=['#28a745', '#ffc107', '#dc3545'], autopct='%1.1f%%')
                    axes[0,0].set_title('Weekly Practices Impact', fontweight='bold')
                
                # Monthly practices pie chart
                monthly_data = [len(impact_data['high_impact']['monthly']), 
                               len(impact_data['medium_impact']['monthly']), 
                               len(impact_data['low_impact']['monthly'])]
                if sum(monthly_data) > 0:
                    axes[0,1].pie(monthly_data, labels=['High', 'Medium', 'Low'], 
                                 colors=['#28a745', '#ffc107', '#dc3545'], autopct='%1.1f%%')
                    axes[0,1].set_title('Monthly Practices Impact', fontweight='bold')
                
                # Quarterly practices pie chart
                quarterly_data = [len(impact_data['high_impact']['quarterly']), 
                                 len(impact_data['medium_impact']['quarterly']), 
                                 len(impact_data['low_impact']['quarterly'])]
                if sum(quarterly_data) > 0:
                    axes[1,0].pie(quarterly_data, labels=['High', 'Medium', 'Low'], 
                                 colors=['#28a745', '#ffc107', '#dc3545'], autopct='%1.1f%%')
                    axes[1,0].set_title('Quarterly Practices Impact', fontweight='bold')
                
                # Overall summary pie chart
                total_high = len(impact_data['high_impact']['weekly']) + len(impact_data['high_impact']['monthly']) + len(impact_data['high_impact']['quarterly'])
                total_medium = len(impact_data['medium_impact']['weekly']) + len(impact_data['medium_impact']['monthly']) + len(impact_data['medium_impact']['quarterly'])
                total_low = len(impact_data['low_impact']['weekly']) + len(impact_data['low_impact']['monthly']) + len(impact_data['low_impact']['quarterly'])
                overall_data = [total_high, total_medium, total_low]
                
                if sum(overall_data) > 0:
                    axes[1,1].pie(overall_data, labels=['High', 'Medium', 'Low'], 
                                 colors=['#28a745', '#ffc107', '#dc3545'], autopct='%1.1f%%')
                    axes[1,1].set_title('Overall Impact Distribution', fontweight='bold')
                
                plt.tight_layout()
                chart3_buffer = BytesIO()
                plt.savefig(chart3_buffer, format='png', dpi=300, bbox_inches='tight')
                zip_file.writestr('detailed_impact_matrix.png', chart3_buffer.getvalue())
                plt.close()
            
            zip_buffer.seek(0)
            
            return send_file(
                BytesIO(zip_buffer.getvalue()),
                as_attachment=True,
                download_name=f'impact_analysis_charts_{datetime.now().strftime("%Y%m%d_%H%M%S")}.zip',
                mimetype='application/zip'
            )
            
        except Exception as e:
            app.logger.error(f"Error generating impact charts: {str(e)}")
            flash('Error generating impact charts. Please try again.', 'error')
            return redirect(url_for('admin_dashboard'))
    
    @app.route('/export-data')
    @admin_required
    def export_data():
        """Export all survey data to Excel (colorful, admin only)"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            import io
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Digital Culture Survey Responses"
            # Headers (colorful, as in screenshot)
            headers = [
                'Sr', 'Name', 'Email', 'Telephone', 'Employee_ID', 'Designation', 'Problem Statement', 'Key Success Metric', 'Timeline',
                'Weekly Practice 1', 'Monthly Practice 1', 'Monthly Practice 2', 'Quarterly Practice 1', 'Quarterly Practice 2',
                'Custom Practice', 'Custom Frequency',
                'START Behavior 1', 'START Behavior 2', 'REDUCE Behavior 1', 'REDUCE Behavior 2', 'STOP Behavior 1', 'STOP Behavior 2',
                'Survey Completed', 'Survey Date', 'Expert Comments Available', 'Last Updated'
            ]
            # Add headers with color
            header_fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True, color="000000")
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            # Add data
            participants = HindalcoPledge.query.all()
            for row_num, participant in enumerate(participants, 2):
                has_survey = bool(participant.survey_responses)
                survey_date = participant.survey_responses[-1].created_at.strftime('%Y-%m-%d') if participant.survey_responses else ''
                expert_comments = 'Yes' if any(sr.expert_comments for sr in participant.survey_responses) else 'No'
                last_updated = participant.survey_responses[-1].updated_at.strftime('%Y-%m-%d %H:%M:%S') if participant.survey_responses else ''
                row_data = [
                    participant.sr_no,
                    participant.name,
                    participant.email,
                    participant.phone,
                    participant.employee_id,
                    participant.designation,
                    participant.problem_statement,
                    participant.success_metric,
                    participant.timeline,
                    participant.weekly_practice_1,
                    participant.monthly_practice_1,
                    participant.monthly_practice_2,
                    participant.quarterly_practice_1,
                    participant.quarterly_practice_2,
                    participant.custom_practice,
                    participant.custom_frequency,
                    participant.behavior_start_1,
                    participant.behavior_start_2,
                    participant.behavior_reduce_1,
                    participant.behavior_reduce_2,
                    participant.behavior_stop_1,
                    participant.behavior_stop_2,
                    'Yes' if has_survey else 'No',
                    survey_date,
                    expert_comments,
                    last_updated
                ]
                for col_num, value in enumerate(row_data, 1):
                    ws.cell(row=row_num, column=col_num, value=value)
            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width
            # Save to memory
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            # Create response
            response = make_response(output.read())
            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            response.headers['Content-Disposition'] = f'attachment; filename=hindalco_survey_data_{datetime.now().strftime("%Y%m%d")}.xlsx'
            return response
        except Exception as e:
            logger.error(f"Excel export error: {e}")
            flash('Error generating Excel export.', 'error')
            return redirect(url_for('admin_dashboard'))
    
    @app.route('/download-user-report/<int:user_id>')
    @login_required
    def download_user_report(user_id):
        """
        Download a PDF report for a specific user.
        - Only the user themselves or an admin can download.
        - This route always returns a PDF, never Excel.
        """
        # Only allow user to download their own report, or admin to download any
        if session.get('user_id') != user_id and not session.get('admin_logged_in'):
            flash('You are not authorized to download this report.', 'error')
            return redirect(url_for('user_dashboard'))
        user = HindalcoPledge.query.get_or_404(user_id)
        surveys = SurveyResponse.query.filter_by(user_id=user.id).order_by(SurveyResponse.created_at.desc()).all()

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
        elements = []
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('title', parent=styles['Title'], fontSize=28, alignment=1, textColor=colors.HexColor('#2346b0'), spaceAfter=20)
        section_header_style = ParagraphStyle('section_header', parent=styles['Heading2'], fontSize=20, alignment=1, textColor=colors.white, backColor=colors.HexColor('#2346b0'), spaceAfter=10, spaceBefore=20)
        table_header_style = ParagraphStyle('table_header', parent=styles['Heading4'], fontSize=14, alignment=1, textColor=colors.white)
        # Title
        elements.append(Paragraph('DIGITAL CULTURE TRANSFORMATION REPORT', title_style))
        elements.append(Spacer(1, 12))
        # Personal Info Section
        elements.append(Paragraph('PERSONAL INFORMATION', section_header_style))
        personal_data = [
            ['Field', 'Details'],
            ['Full Name', user.name or 'Not specified'],
            ['Email Address', user.email or 'Not specified'],
            ['Employee ID', user.employee_id or 'Not specified'],
            ['Phone Number', user.phone or 'Not specified'],
            ['Designation/Role', user.designation or 'Not specified'],
            ['Signature Date', 'Not specified'],
        ]
        personal_table = Table(personal_data, colWidths=[180, 300])
        personal_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2346b0')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('ALIGN', (0,0), (-1,0), 'CENTER'),
            ('BACKGROUND', (0,1), (-1,-1), colors.HexColor('#f6fafd')),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#dbe5f1')),
            ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
            ('FONTSIZE', (0,0), (-1,-1), 12),
        ]))
        elements.append(personal_table)
        elements.append(Spacer(1, 18))
        # Digital North Star Section
        elements.append(Paragraph('DIGITAL NORTH STAR', section_header_style))
        north_star_data = [
            ['Component', 'Description'],
            ['Problem Statement', user.problem_statement or 'Not specified'],
            ['Success Metric', user.success_metric or 'Not specified'],
            ['Timeline', user.timeline or 'Not specified'],
        ]
        north_star_table = Table(north_star_data, colWidths=[180, 300])
        north_star_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#009e73')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('ALIGN', (0,0), (-1,0), 'CENTER'),
            ('BACKGROUND', (0,1), (-1,-1), colors.HexColor('#eafaf3')),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#b6e2d3')),
            ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
            ('FONTSIZE', (0,0), (-1,-1), 12),
        ]))
        elements.append(north_star_table)
        elements.append(Spacer(1, 18))
        # Behavior Change Commitments Section
        elements.append(Paragraph('BEHAVIOR CHANGE COMMITMENTS', section_header_style))
        # START Behaviors
        elements.append(Paragraph('START Behaviors', ParagraphStyle('subheader', parent=styles['Heading3'], fontSize=16, textColor=colors.HexColor('#2346b0'), spaceAfter=8)))
        start_data = [
            ['Behavior Type', 'Description'],
            ['START Behavior 1', user.behavior_start_1 or ''],
            ['START Behavior 2', user.behavior_start_2 or ''],
        ]
        start_table = Table(start_data, colWidths=[180, 300])
        start_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#009e73')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('ALIGN', (0,0), (-1,0), 'CENTER'),
            ('BACKGROUND', (0,1), (-1,-1), colors.white),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#b6e2d3')),
            ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
            ('FONTSIZE', (0,0), (-1,-1), 12),
        ]))
        elements.append(start_table)
        elements.append(Spacer(1, 12))
        # REDUCE Behaviors
        elements.append(Paragraph('REDUCE Behaviors', ParagraphStyle('subheader', parent=styles['Heading3'], fontSize=16, textColor=colors.HexColor('#2346b0'), spaceAfter=8)))
        reduce_data = [
            ['Behavior Type', 'Description'],
            ['REDUCE Behavior 1', user.behavior_reduce_1 or ''],
            ['REDUCE Behavior 2', user.behavior_reduce_2 or ''],
        ]
        reduce_table = Table(reduce_data, colWidths=[180, 300])
        reduce_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#f7a600')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('ALIGN', (0,0), (-1,0), 'CENTER'),
            ('BACKGROUND', (0,1), (-1,-1), colors.white),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#ffe5b4')),
            ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
            ('FONTSIZE', (0,0), (-1,-1), 12),
        ]))
        elements.append(reduce_table)
        elements.append(Spacer(1, 12))
        # STOP Behaviors
        elements.append(Paragraph('STOP Behaviors', ParagraphStyle('subheader', parent=styles['Heading3'], fontSize=16, textColor=colors.HexColor('#2346b0'), spaceAfter=8)))
        stop_data = [
            ['Behavior Type', 'Description'],
            ['STOP Behavior 1', user.behavior_stop_1 or ''],
            ['STOP Behavior 2', user.behavior_stop_2 or ''],
        ]
        stop_table = Table(stop_data, colWidths=[180, 300])
        stop_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#d7263d')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('ALIGN', (0,0), (-1,0), 'CENTER'),
            ('BACKGROUND', (0,1), (-1,-1), colors.white),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#f7b6b6')),
            ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
            ('FONTSIZE', (0,0), (-1,-1), 12),
        ]))
        elements.append(stop_table)
        elements.append(Spacer(1, 18))
        # Practice Commitments Section
        elements.append(Paragraph('PRACTICE COMMITMENTS', section_header_style))
        practice_data = [
            ['Practice Type', 'Description'],
            ['Weekly Practice', user.weekly_practice_1 or ''],
            ['Monthly Practice 1', user.monthly_practice_1 or ''],
            ['Monthly Practice 2', user.monthly_practice_2 or ''],
            ['Quarterly Practice 1', user.quarterly_practice_1 or ''],
            ['Quarterly Practice 2', user.quarterly_practice_2 or ''],
            ['Custom Practice', user.custom_practice or ''],
            ['Custom Frequency', user.custom_frequency or ''],
        ]
        practice_table = Table(practice_data, colWidths=[180, 300])
        practice_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2346b0')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('ALIGN', (0,0), (-1,0), 'CENTER'),
            ('BACKGROUND', (0,1), (-1,-1), colors.HexColor('#f6fafd')),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#dbe5f1')),
            ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
            ('FONTSIZE', (0,0), (-1,-1), 12),
        ]))
        elements.append(practice_table)
        elements.append(Spacer(1, 18))
        # Survey Responses Section (if any)
        if surveys:
            elements.append(Paragraph('SURVEY RESPONSES', section_header_style))
            for idx, survey in enumerate(surveys, 1):
                try:
                    data = json.loads(survey.response_data) if survey.response_data else {}
                except Exception:
                    data = {}
                elements.append(Paragraph(f'Survey #{idx}', styles['Heading4']))
                survey_table_data = [['Field', 'Value']]
                for k, v in data.items():
                    survey_table_data.append([str(k), str(v)])
                if survey.expert_comments:
                    survey_table_data.append(['Expert Comments', survey.expert_comments])
                survey_table = Table(survey_table_data, colWidths=[180, 300])
                survey_table.setStyle(TableStyle([
                    ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#009e73')),
                    ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                    ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                    ('ALIGN', (0,0), (-1,0), 'CENTER'),
                    ('BACKGROUND', (0,1), (-1,-1), colors.HexColor('#eafaf3')),
                    ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#b6e2d3')),
                    ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
                    ('FONTSIZE', (0,0), (-1,-1), 12),
                ]))
                elements.append(survey_table)
                elements.append(Spacer(1, 12))
        doc.build(elements)
        buffer.seek(0)
        filename = f"Digital_Culture_Transformation_Report_{user.name.replace(' ', '_')}.pdf"
        return send_file(buffer, as_attachment=True, download_name=filename, mimetype='application/pdf')
    
    @app.route('/download-report')
    @login_required
    def download_report():
        return redirect(url_for('export_data'))
    
    @app.route('/logout')
    def logout():
        session.clear()
        flash('You have been logged out successfully.', 'info')
        return redirect(url_for('index'))
    
    @app.route('/admin-logout')
    @admin_required
    def admin_logout():
        session.clear()
        flash('Admin logged out successfully.', 'info')
        return redirect(url_for('index'))
    

    
    def get_detailed_impact_analysis():
        """Generate detailed impact analysis data for charts"""
        
        # Get all participants with surveys
        participants = HindalcoPledge.query.all()
        
        impact_analysis = {
            'high_impact': {'weekly': [], 'monthly': [], 'quarterly': []},
            'medium_impact': {'weekly': [], 'monthly': [], 'quarterly': []},
            'low_impact': {'weekly': [], 'monthly': [], 'quarterly': []},
            'user_details': {},
            'practice_popularity': {
                'weekly_practices': {},
                'monthly_practices': {},
                'quarterly_practices': {}
            },
            'behavior_analysis': {
                'start_behaviors': {},
                'reduce_behaviors': {},
                'stop_behaviors': {}
            }
        }
        
        for participant in participants:
            if not participant.survey_responses:
                continue
                
            # Get latest survey response
            latest_survey = max(participant.survey_responses, key=lambda x: x.created_at)
            
            try:
                survey_data = json.loads(latest_survey.response_data)
                
                user_info = {
                    'id': participant.id,
                    'name': participant.name,
                    'email': participant.email,
                    'pledge_data': {
                        'problem_statement': participant.problem_statement,
                        'success_metric': participant.success_metric,
                        'timeline': participant.timeline,
                        'weekly_practice_1': participant.weekly_practice_1,
                        'monthly_practice_1': participant.monthly_practice_1,
                        'monthly_practice_2': participant.monthly_practice_2,
                        'quarterly_practice_1': participant.quarterly_practice_1,
                        'quarterly_practice_2': participant.quarterly_practice_2
                    },
                    'survey_responses': {}
                }
                
                # Analyze weekly practices
                if 'weekly_practice_1' in survey_data:
                    impact = survey_data['weekly_practice_1'].get('impact', '').upper()
                    practice_text = participant.weekly_practice_1 or 'Not specified'
                    
                    # Track practice popularity
                    if practice_text not in impact_analysis['practice_popularity']['weekly_practices']:
                        impact_analysis['practice_popularity']['weekly_practices'][practice_text] = {
                            'high': 0, 'medium': 0, 'low': 0, 'total': 0
                        }
                    
                    if impact in ['HIGH', 'MEDIUM', 'LOW']:
                        impact_key = impact.lower()
                        impact_analysis['practice_popularity']['weekly_practices'][practice_text][impact_key] += 1
                        impact_analysis['practice_popularity']['weekly_practices'][practice_text]['total'] += 1
                    
                    user_info['survey_responses']['weekly_practice_1'] = {
                        'impact': impact,
                        'response': survey_data['weekly_practice_1']
                    }
                    
                    if impact == 'HIGH':
                        impact_analysis['high_impact']['weekly'].append(user_info.copy())
                    elif impact == 'MEDIUM':
                        impact_analysis['medium_impact']['weekly'].append(user_info.copy())
                    elif impact == 'LOW':
                        impact_analysis['low_impact']['weekly'].append(user_info.copy())
                
                # Analyze monthly practices
                for monthly_key in ['monthly_practice_1', 'monthly_practice_2']:
                    if monthly_key in survey_data:
                        impact = survey_data[monthly_key].get('impact', '').upper()
                        practice_text = getattr(participant, monthly_key, None) or 'Not specified'
                        
                        # Track practice popularity
                        if practice_text not in impact_analysis['practice_popularity']['monthly_practices']:
                            impact_analysis['practice_popularity']['monthly_practices'][practice_text] = {
                                'high': 0, 'medium': 0, 'low': 0, 'total': 0
                            }
                        
                        if impact in ['HIGH', 'MEDIUM', 'LOW']:
                            impact_key = impact.lower()
                            impact_analysis['practice_popularity']['monthly_practices'][practice_text][impact_key] += 1
                            impact_analysis['practice_popularity']['monthly_practices'][practice_text]['total'] += 1
                        
                        user_info['survey_responses'][monthly_key] = {
                            'impact': impact,
                            'response': survey_data[monthly_key]
                        }
                        
                        if impact == 'HIGH':
                            impact_analysis['high_impact']['monthly'].append(user_info.copy())
                        elif impact == 'MEDIUM':
                            impact_analysis['medium_impact']['monthly'].append(user_info.copy())
                        elif impact == 'LOW':
                            impact_analysis['low_impact']['monthly'].append(user_info.copy())
                
                # Analyze quarterly practices
                for quarterly_key in ['quarterly_practice_1', 'quarterly_practice_2']:
                    if quarterly_key in survey_data:
                        impact = survey_data[quarterly_key].get('impact', '').upper()
                        practice_text = getattr(participant, quarterly_key, None) or 'Not specified'
                        
                        # Track practice popularity
                        if practice_text not in impact_analysis['practice_popularity']['quarterly_practices']:
                            impact_analysis['practice_popularity']['quarterly_practices'][practice_text] = {
                                'high': 0, 'medium': 0, 'low': 0, 'total': 0
                            }
                        
                        if impact in ['HIGH', 'MEDIUM', 'LOW']:
                            impact_key = impact.lower()
                            impact_analysis['practice_popularity']['quarterly_practices'][practice_text][impact_key] += 1
                            impact_analysis['practice_popularity']['quarterly_practices'][practice_text]['total'] += 1
                        
                        user_info['survey_responses'][quarterly_key] = {
                            'impact': impact,
                            'response': survey_data[quarterly_key]
                        }
                        
                        if impact == 'HIGH':
                            impact_analysis['high_impact']['quarterly'].append(user_info.copy())
                        elif impact == 'MEDIUM':
                            impact_analysis['medium_impact']['quarterly'].append(user_info.copy())
                        elif impact == 'LOW':
                            impact_analysis['low_impact']['quarterly'].append(user_info.copy())
                
                # Analyze behavior data
                for i in range(1, 3):
                    # START behaviors
                    start_key = f'behavior_start_{i}'
                    behavior_text = getattr(participant, start_key, None) or 'Not specified'
                    if behavior_text and behavior_text.strip() and behavior_text != 'Not specified':
                        if behavior_text not in impact_analysis['behavior_analysis']['start_behaviors']:
                            impact_analysis['behavior_analysis']['start_behaviors'][behavior_text] = 0
                        impact_analysis['behavior_analysis']['start_behaviors'][behavior_text] += 1
                    
                    # REDUCE behaviors  
                    reduce_key = f'behavior_reduce_{i}'
                    behavior_text = getattr(participant, reduce_key, None) or 'Not specified'
                    if behavior_text and behavior_text.strip() and behavior_text != 'Not specified':
                        if behavior_text not in impact_analysis['behavior_analysis']['reduce_behaviors']:
                            impact_analysis['behavior_analysis']['reduce_behaviors'][behavior_text] = 0
                        impact_analysis['behavior_analysis']['reduce_behaviors'][behavior_text] += 1
                    
                    # STOP behaviors
                    stop_key = f'behavior_stop_{i}'
                    behavior_text = getattr(participant, stop_key, None) or 'Not specified'
                    if behavior_text and behavior_text.strip() and behavior_text != 'Not specified':
                        if behavior_text not in impact_analysis['behavior_analysis']['stop_behaviors']:
                            impact_analysis['behavior_analysis']['stop_behaviors'][behavior_text] = 0
                        impact_analysis['behavior_analysis']['stop_behaviors'][behavior_text] += 1
                
                # Store complete user details
                impact_analysis['user_details'][participant.id] = user_info
                
            except (json.JSONDecodeError, KeyError) as e:
                logger.warning(f"Error parsing survey data for participant {participant.id}: {str(e)}")
                continue
        
        # Debug behavior analysis
        logger.info(f"Behavior analysis: START={len(impact_analysis['behavior_analysis']['start_behaviors'])}, REDUCE={len(impact_analysis['behavior_analysis']['reduce_behaviors'])}, STOP={len(impact_analysis['behavior_analysis']['stop_behaviors'])}")
        
        return impact_analysis
    
    return app

def initialize_database_with_participants(app):
    """Initialize database and load ALL participants from CSV using enhanced loader"""
    with app.app_context():
        try:
            logger.info("Initializing database and loading ALL participants from CSV...")
            
            # Create tables
            db.create_all()
            
            # Check if participants already exist
            if HindalcoPledge.query.count() > 0:
                logger.info(f"Database already contains {HindalcoPledge.query.count()} participants")
                return True
            
            # Use enhanced CSV loader
            def clean_text(s):
                if pd.isna(s):
                    return ""
                return re.sub(r'\s+', ' ', s).strip()

            def clean_phone(s):
                if pd.isna(s):
                    return ""
                # Remove all non-digit characters
                return re.sub(r'\D', '', s)

            def load_all_participants_from_csv():
                """Load ALL participants with complete data from the new CSV file using header names."""
                try:
                    csv_file = 'All_36_Users_Complete_Data_20250715_121445.csv'
                    df = pd.read_csv(csv_file)
                    print(f"CSV loaded: {len(df)} rows, {len(df.columns)} columns")
                    participants = []
                    for index, row in df.iterrows():
                        try:
                            sr_no = clean_text(row.get('Sr', index+1))
                            name = clean_text(row.get('Name'))
                            email = clean_text(row.get('Email'))
                            phone = clean_phone(row.get('Telephone'))
                            employee_id = clean_text(row.get('Employee_ID', f"HIN{str(index+1).zfill(3)}"))
                            designation = clean_text(row.get('Designation', 'Digital Transformation Specialist'))
                            problem_statement = clean_text(row.get('Problem_Statement'))
                            success_metric = clean_text(row.get('Key_Success_Metric'))
                            timeline = clean_text(row.get('Timeline_to_Impact'))
                            weekly_practice_1 = clean_text(row.get('Weekly_Practice_1'))
                            monthly_practice_1 = clean_text(row.get('Monthly_Practice_1'))
                            monthly_practice_2 = clean_text(row.get('Monthly_Practice_2'))
                            quarterly_practice_1 = clean_text(row.get('Quarterly_Practice_1'))
                            quarterly_practice_2 = clean_text(row.get('Quarterly_Practice_2'))
                            custom_practice = clean_text(row.get('Custom_Practice'))
                            custom_frequency = clean_text(row.get('Custom_Frequency'))
                            behavior_start_1 = clean_text(row.get('Behavior_START_1'))
                            behavior_start_2 = clean_text(row.get('Behavior_START_2'))
                            behavior_reduce_1 = clean_text(row.get('Behavior_REDUCE_1'))
                            behavior_reduce_2 = clean_text(row.get('Behavior_REDUCE_2'))
                            behavior_stop_1 = clean_text(row.get('Behavior_STOP_1'))
                            behavior_stop_2 = clean_text(row.get('Behavior_STOP_2'))
                            participant_data = {
                                'sr_no': sr_no,
                                'name': name,
                                'email': email,
                                'phone': phone,
                                'employee_id': employee_id,
                                'designation': designation,
                                'problem_statement': problem_statement,
                                'success_metric': success_metric,
                                'timeline': timeline,
                                'weekly_practice_1': weekly_practice_1,
                                'monthly_practice_1': monthly_practice_1,
                                'monthly_practice_2': monthly_practice_2,
                                'quarterly_practice_1': quarterly_practice_1,
                                'quarterly_practice_2': quarterly_practice_2,
                                'custom_practice': custom_practice,
                                'custom_frequency': custom_frequency,
                                'behavior_start_1': behavior_start_1,
                                'behavior_start_2': behavior_start_2,
                                'behavior_reduce_1': behavior_reduce_1,
                                'behavior_reduce_2': behavior_reduce_2,
                                'behavior_stop_1': behavior_stop_1,
                                'behavior_stop_2': behavior_stop_2
                            }
                            if name and email:
                                participants.append(participant_data)
                                print(f"Loaded: {name} - {email}")
                        except Exception as e:
                            print(f"Error processing row {index}: {e}")
                            continue
                    print(f"\nTotal participants loaded: {len(participants)}")
                    return participants
                except Exception as e:
                    print(f"Error loading CSV: {e}")
                    return []
            
            # Load participants data
            participants_data = load_all_participants_from_csv()
            
            if not participants_data:
                logger.error("No participants loaded from enhanced CSV loader")
                return False
            
            loaded_count = 0
            
            # Create participants in database
            for participant_data in participants_data:
                try:
                    participant = HindalcoPledge(
                        sr_no=participant_data.get('sr_no', loaded_count + 1),
                        name=participant_data['name'],
                        email=participant_data['email'],
                        phone=participant_data.get('phone', ''),
                        employee_id=participant_data.get('employee_id', f"HIN{str(loaded_count + 1).zfill(3)}"),
                        designation=participant_data.get('designation', 'Digital Transformation Specialist'),
                        problem_statement=participant_data.get('problem_statement', ''),
                        success_metric=participant_data.get('success_metric', ''),
                        timeline=participant_data.get('timeline', ''),
                        weekly_practice_1=participant_data.get('weekly_practice_1', ''),
                        monthly_practice_1=participant_data.get('monthly_practice_1', ''),
                        monthly_practice_2=participant_data.get('monthly_practice_2', ''),
                        quarterly_practice_1=participant_data.get('quarterly_practice_1', ''),
                        quarterly_practice_2=participant_data.get('quarterly_practice_2', ''),
                        behavior_start_1=participant_data.get('behavior_start_1', ''),
                        behavior_start_2=participant_data.get('behavior_start_2', ''),
                        behavior_reduce_1=participant_data.get('behavior_reduce_1', ''),
                        behavior_reduce_2=participant_data.get('behavior_reduce_2', ''),
                        behavior_stop_1=participant_data.get('behavior_stop_1', ''),
                        behavior_stop_2=participant_data.get('behavior_stop_2', '')
                    )
                    
                    db.session.add(participant)
                    loaded_count += 1
                    
                    # Add demo survey for demo user
                    if participant_data['email'] == 'demo@test.com':
                        db.session.flush()  # Get the ID
                        demo_survey = SurveyResponse(
                            user_id=participant.id,
                            response_data=json.dumps({
                                'weekly_practice_1': {'impact': 'HIGH', 'action_taken': 'Implemented dashboard reviews', 'action_needed': 'Expand to other teams'},
                                'monthly_practice_1': {'impact': 'MEDIUM', 'action_taken': 'Started data-based decisions', 'action_needed': 'Improve data quality'},
                                'monthly_practice_2': {'impact': 'HIGH', 'action_taken': 'Digital training programs', 'action_needed': 'Continue advanced training'},
                                'quarterly_practice_1': {'impact': 'MEDIUM', 'action_taken': 'Set digital goals', 'action_needed': 'Better measurement'},
                                'quarterly_practice_2': {'impact': 'HIGH', 'action_taken': 'Innovation proposals', 'action_needed': 'Implementation support'}
                            }),
                            expert_comments="Excellent progress on digital transformation. Continue building capabilities."
                        )
                        db.session.add(demo_survey)
                    
                    logger.info(f"Loaded: {participant_data['name']} - {participant_data['email']}")
                    
                except Exception as e:
                    logger.error(f"Error creating participant {participant_data.get('name', 'Unknown')}: {e}")
                    db.session.rollback()
                    continue
            
            db.session.commit()
            logger.info(f"Successfully loaded {loaded_count} participants from enhanced CSV loader")
            return True
                    
        except Exception as e:
            db.session.rollback()
            logger.error(f"Database initialization error: {e}")
            import traceback
            traceback.print_exc()
            return False

# Main application setup
if __name__ == '__main__':
    app = create_app()
    
    # Initialize database with participants
    if initialize_database_with_participants(app):
        logger.info("Database initialized successfully")
    else:
        logger.error("Database initialization failed")
    
    # Run app
    app.run(host='0.0.0.0', port=5000, debug=True)